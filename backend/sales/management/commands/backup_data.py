"""
Management command to backup all database data to Excel and JSON files
Usage: python manage.py backup_data
"""
from django.core.management.base import BaseCommand
from django.core.serializers import serialize
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sales.models import Customer, DailyRate, Purchase, Sale, Payment, Expense
import json
import os


class Command(BaseCommand):
    help = 'Backup all database data to Excel and JSON files'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            type=str,
            default='backups',
            help='Directory to store backup files (default: backups/)'
        )

    def handle(self, *args, **options):
        output_dir = options['output_dir']
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        self.stdout.write(self.style.SUCCESS(f'\n🔄 Starting backup process...'))
        self.stdout.write(f'📁 Output directory: {output_dir}')
        self.stdout.write(f'⏰ Timestamp: {timestamp}\n')

        # Backup to Excel
        excel_file = self.backup_to_excel(output_dir, timestamp)
        
        # Backup to JSON (for complete data restoration)
        json_file = self.backup_to_json(output_dir, timestamp)
        
        self.stdout.write(self.style.SUCCESS(f'\n✅ Backup completed successfully!'))
        self.stdout.write(f'📊 Excel file: {excel_file}')
        self.stdout.write(f'📄 JSON file: {json_file}')
        self.stdout.write(f'💾 Backup timestamp: {timestamp}\n')

    def backup_to_excel(self, output_dir, timestamp):
        """Create Excel backup with all data"""
        filename = os.path.join(output_dir, f'ahmad_poultry_backup_{timestamp}.xlsx')
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Backup each model
        self._backup_customers(wb)
        self._backup_daily_rates(wb)
        self._backup_purchases(wb)
        self._backup_sales(wb)
        self._backup_payments(wb)
        self._backup_expenses(wb)
        
        # Add summary sheet
        self._add_summary_sheet(wb, timestamp)
        
        wb.save(filename)
        return filename

    def _style_header(self, ws):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _backup_customers(self, wb):
        """Backup customers to Excel sheet"""
        ws = wb.create_sheet("Customers")
        
        # Headers
        headers = ['ID', 'Name', 'Phone', 'Address', 'Opening Balance', 
                  'Running Balance', 'Is Active', 'Created At', 'Updated At']
        ws.append(headers)
        
        # Data
        customers = Customer.objects.all().order_by('id')
        for customer in customers:
            ws.append([
                customer.id,
                customer.name,
                customer.phone,
                customer.address,
                float(customer.opening_balance),
                float(customer.running_balance),
                'Yes' if customer.is_active else 'No',
                customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                customer.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self._style_header(ws)
        self.stdout.write(f'✓ Backed up {customers.count()} customers')

    def _backup_daily_rates(self, wb):
        """Backup daily rates to Excel sheet"""
        ws = wb.create_sheet("Daily Rates")
        
        # Headers
        headers = ['ID', 'Date', 'Default Cost Rate', 'Default Sale Rate', 
                  'Created At', 'Updated At']
        ws.append(headers)
        
        # Data
        rates = DailyRate.objects.all().order_by('-date')
        for rate in rates:
            ws.append([
                rate.id,
                rate.date.strftime('%Y-%m-%d'),
                float(rate.default_cost_rate),
                float(rate.default_sale_rate),
                rate.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                rate.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self._style_header(ws)
        self.stdout.write(f'✓ Backed up {rates.count()} daily rates')

    def _backup_purchases(self, wb):
        """Backup purchases to Excel sheet"""
        ws = wb.create_sheet("Purchases")
        
        # Headers
        headers = ['ID', 'Date', 'Supplier', 'KG', 'Cost Rate per KG', 
                  'Total Cost', 'Note', 'Created At', 'Updated At']
        ws.append(headers)
        
        # Data
        purchases = Purchase.objects.all().order_by('-date', '-created_at')
        for purchase in purchases:
            ws.append([
                purchase.id,
                purchase.date.strftime('%Y-%m-%d'),
                purchase.supplier,
                float(purchase.kg),
                float(purchase.cost_rate_per_kg),
                float(purchase.total_cost),
                purchase.note,
                purchase.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                purchase.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self._style_header(ws)
        self.stdout.write(f'✓ Backed up {purchases.count()} purchases')

    def _backup_sales(self, wb):
        """Backup sales to Excel sheet"""
        ws = wb.create_sheet("Sales")
        
        # Headers
        headers = ['ID', 'Date', 'Customer', 'KG', 'Sale Rate per KG', 
                  'Cost Rate Snapshot', 'Total Amount', 'Amount Received', 
                  'Borrow Amount', 'Profit', 'Note', 'Created At', 'Updated At']
        ws.append(headers)
        
        # Data
        sales = Sale.objects.select_related('customer').all().order_by('-date', '-created_at')
        for sale in sales:
            ws.append([
                sale.id,
                sale.date.strftime('%Y-%m-%d'),
                sale.customer.name,
                float(sale.kg),
                float(sale.sale_rate_per_kg),
                float(sale.cost_rate_snapshot),
                float(sale.total_amount),
                float(sale.amount_received),
                float(sale.borrow_amount),
                float(sale.profit),
                sale.note,
                sale.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                sale.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self._style_header(ws)
        self.stdout.write(f'✓ Backed up {sales.count()} sales')

    def _backup_payments(self, wb):
        """Backup payments to Excel sheet"""
        ws = wb.create_sheet("Payments")
        
        # Headers
        headers = ['ID', 'Date', 'Customer', 'Amount', 'Payment Method', 
                  'Note', 'Created At', 'Updated At']
        ws.append(headers)
        
        # Data
        payments = Payment.objects.select_related('customer').all().order_by('-date', '-created_at')
        for payment in payments:
            ws.append([
                payment.id,
                payment.date.strftime('%Y-%m-%d'),
                payment.customer.name,
                float(payment.amount),
                payment.get_method_display(),
                payment.note,
                payment.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                payment.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self._style_header(ws)
        self.stdout.write(f'✓ Backed up {payments.count()} payments')

    def _backup_expenses(self, wb):
        """Backup expenses to Excel sheet"""
        ws = wb.create_sheet("Expenses")
        
        # Headers
        headers = ['ID', 'Date', 'Category', 'Amount', 'Note', 
                  'Created At', 'Updated At']
        ws.append(headers)
        
        # Data
        expenses = Expense.objects.all().order_by('-date', '-created_at')
        for expense in expenses:
            ws.append([
                expense.id,
                expense.date.strftime('%Y-%m-%d'),
                expense.get_category_display(),
                float(expense.amount),
                expense.note,
                expense.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                expense.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self._style_header(ws)
        self.stdout.write(f'✓ Backed up {expenses.count()} expenses')

    def _add_summary_sheet(self, wb, timestamp):
        """Add summary sheet with backup information"""
        ws = wb.create_sheet("Backup Summary", 0)  # Make it first sheet
        
        # Title
        ws['A1'] = 'Ahmad Poultry Services - Data Backup'
        ws['A1'].font = Font(bold=True, size=16, color="4472C4")
        ws.merge_cells('A1:B1')
        
        # Backup info
        ws['A3'] = 'Backup Date & Time:'
        ws['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Backup Timestamp:'
        ws['B4'] = timestamp
        
        # Record counts
        ws['A6'] = 'Data Summary:'
        ws['A6'].font = Font(bold=True, size=12)
        
        ws['A7'] = 'Customers:'
        ws['B7'] = Customer.objects.count()
        
        ws['A8'] = 'Daily Rates:'
        ws['B8'] = DailyRate.objects.count()
        
        ws['A9'] = 'Purchases:'
        ws['B9'] = Purchase.objects.count()
        
        ws['A10'] = 'Sales:'
        ws['B10'] = Sale.objects.count()
        
        ws['A11'] = 'Payments:'
        ws['B11'] = Payment.objects.count()
        
        ws['A12'] = 'Expenses:'
        ws['B12'] = Expense.objects.count()
        
        ws['A14'] = 'Total Records:'
        ws['B14'] = (Customer.objects.count() + DailyRate.objects.count() + 
                     Purchase.objects.count() + Sale.objects.count() + 
                     Payment.objects.count() + Expense.objects.count())
        ws['A14'].font = Font(bold=True)
        ws['B14'].font = Font(bold=True)
        
        # Notes
        ws['A16'] = 'Notes:'
        ws['A16'].font = Font(bold=True)
        ws['A17'] = '• This backup includes all data from the database'
        ws['A18'] = '• Each sheet contains data from one model/table'
        ws['A19'] = '• Running Balance is calculated at backup time'
        ws['A20'] = '• For data restoration, use the JSON backup file'
        
        # Style
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30

    def backup_to_json(self, output_dir, timestamp):
        """Create JSON backup with complete data for restoration"""
        filename = os.path.join(output_dir, f'ahmad_poultry_backup_{timestamp}.json')
        
        backup_data = {
            'backup_timestamp': timestamp,
            'backup_datetime': timezone.now().isoformat(),
            'data': {
                'customers': json.loads(serialize('json', Customer.objects.all())),
                'daily_rates': json.loads(serialize('json', DailyRate.objects.all())),
                'purchases': json.loads(serialize('json', Purchase.objects.all())),
                'sales': json.loads(serialize('json', Sale.objects.all())),
                'payments': json.loads(serialize('json', Payment.objects.all())),
                'expenses': json.loads(serialize('json', Expense.objects.all())),
            },
            'counts': {
                'customers': Customer.objects.count(),
                'daily_rates': DailyRate.objects.count(),
                'purchases': Purchase.objects.count(),
                'sales': Sale.objects.count(),
                'payments': Payment.objects.count(),
                'expenses': Expense.objects.count(),
            }
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, indent=2, ensure_ascii=False)
        
        self.stdout.write(f'✓ Created JSON backup with {sum(backup_data["counts"].values())} total records')
        
        return filename

